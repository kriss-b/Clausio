"""Import du clausier du Club RSSI Santé / Club DPO / AFIB et lecture des fiches de réponse.

La fiche « SynAApCE » (onglet RetourClausier) est la version machine-lisible du
clausier : chaque exigence porte son code d'article (O-x.y). On l'exploite deux fois :

  1) IMPORT : générer le référentiel Clausio à partir des vrais codes d'articles
     (pas d'une recopie manuelle) — maintenable quand le club publie une nouvelle version.
  2) LECTURE : quand un candidat retourne la fiche remplie, lire sa réponse STRUCTURÉE
     par exigence (Mesures Oui/Non + Précisions + Éléments de consolidation), bien
     supérieure à l'extraction en texte libre d'un PDF.

Le champ « Mesures Oui/Non » est l'auto-appréciation du candidat : Clausio ne la prend
jamais pour argent comptant, il la confronte aux précisions et aux critères.
"""
from __future__ import annotations

import re
import warnings
from pathlib import Path

warnings.filterwarnings("ignore")

_ART = re.compile(r"^[O0][-\s]?(\d+(?:\.\d+)*)$")   # O-3.1, O-12.2.3.2, tolère 0-2.5
_CHAP = re.compile(r"^\d+(?:\.\d+)?$")

# chapitre -> (axe, profils, criticité par défaut, conditionnel)
CHAPITRES = {
    "1":    ("gouvernance",                  ["socle", "dispositif_medical"], "mineur", False),
    "2":    ("sous_traitance",               ["socle", "dispositif_medical"], "majeur", True),
    "3":    ("logiciels",                    ["socle", "dispositif_medical"], "majeur", False),
    "4":    ("identites",                    ["socle", "dispositif_medical"], "majeur", False),
    "5":    ("authentification",             ["socle", "dispositif_medical"], "majeur", False),
    "6":    ("tracabilite",                  ["socle", "dispositif_medical"], "majeur", False),
    "7":    ("protection_systemes",          ["socle", "dispositif_medical"], "majeur", False),
    "8":    ("cryptographie",                ["socle", "dispositif_medical"], "majeur", False),
    "9":    ("maintenance",                  ["socle", "dispositif_medical"], "majeur", True),
    "10":   ("reseau_wifi",                  ["socle", "dispositif_medical"], "mineur", True),
    "11":   ("protection_donnees_medicales", ["socle", "dispositif_medical"], "majeur", True),
    "12.1": ("dispositifs_mobiles",          ["socle", "dispositif_medical"], "mineur", True),
    "12.2": ("dispositif_medical_connecte",  ["dispositif_medical"],          "majeur", True),
    "12.3": ("hebergement",                  ["socle", "dispositif_medical"], "majeur", True),
    "12.4": ("hebergement",                  ["socle", "dispositif_medical"], "majeur", True),
    "12.5": ("developpement",                ["socle", "dispositif_medical"], "mineur", True),
    "13":   ("intelligence_artificielle",    ["socle", "dispositif_medical", "ia"], "mineur", True),
}

# Criticité renforcée pour quelques points structurants (proposition — le RSSI affine).
BLOQUANTS = {"O-2.5"}  # notification de violation de données sous 24 h


def _chapitre_du_code(code: str) -> str:
    """O-12.2.3.2 -> '12.2' ; O-3.1 -> '3'."""
    parts = code[2:].split(".")
    if parts[0] == "12" and len(parts) >= 2:
        return f"12.{parts[1]}"
    return parts[0]


def _cellule(row, i: int) -> str:
    return "" if (len(row) <= i or row[i] is None) else str(row[i]).strip()


def _ouvrir_retour_clausier(path: Path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    if "RetourClausier" not in wb.sheetnames:
        return None
    return list(wb["RetourClausier"].iter_rows(values_only=True))


def importer_referentiel(path: Path) -> list[dict]:
    """Retourne la liste des exigences (dicts prêts pour le YAML/la base)."""
    rows = _ouvrir_retour_clausier(path)
    if rows is None:
        raise ValueError("Onglet 'RetourClausier' introuvable : ce n'est pas une fiche SynAApCE.")

    exigences: list[dict] = []
    for r in rows[1:]:
        brut = _cellule(r, 0)
        m = _ART.match(brut)
        if not m:
            continue
        code = "O-" + m.group(1)                 # normalise 0- -> O-
        chap = _chapitre_du_code(code)
        axe, profils, crit, conditionnel = CHAPITRES.get(chap, ("autre", ["socle"], "majeur", False))
        descriptif = _cellule(r, 1).replace("\n", " ").strip()
        consolidation = _cellule(r, 4).replace("\n", " ").strip()

        exigences.append({
            "code": code,
            "chapitre": chap,
            "axe": axe,
            "libelle": descriptif,
            "profils": profils,
            "criticite_defaut": "bloquant" if code in BLOQUANTS else crit,
            "conditionnel": conditionnel,
            "question_rag": descriptif,
            "criteres_acceptation": (
                [f"Réponse 'Oui' étayée par des précisions concrètes.",
                 f"Éléments attendus : {consolidation}"] if consolidation
                else ["Réponse 'Oui' étayée par des précisions concrètes et vérifiables."]
            ),
        })
    return exigences


def lire_fiche_reponse(path: Path) -> dict[str, dict]:
    """Fiche remplie -> {code_exigence: {mesure, precisions, consolidation}}."""
    rows = _ouvrir_retour_clausier(path)
    if rows is None:
        return {}
    reponses: dict[str, dict] = {}
    for r in rows[1:]:
        m = _ART.match(_cellule(r, 0))
        if not m:
            continue
        code = "O-" + m.group(1)
        reponses[code] = {
            "mesure": _cellule(r, 2),          # Oui / Non
            "precisions": _cellule(r, 3),      # précisions du candidat
            "consolidation": _cellule(r, 4),   # éléments de consolidation
        }
    return reponses


def est_fiche_reponse(path: Path) -> bool:
    if path.suffix.lower() != ".xlsx":
        return False
    try:
        return _ouvrir_retour_clausier(path) is not None
    except Exception:  # noqa: BLE001
        return False


# --- Fichier de liaison Clausio (rapport -> candidat -> réimport pour reprise) ---
FEUILLE_ID = "Clausio_Identification"
FEUILLE_DEMANDES = "Demandes"


def est_fiche_liaison(path: Path) -> bool:
    if path.suffix.lower() != ".xlsx":
        return False
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)   # non read_only : gère les validations étendues d'Excel
        return FEUILLE_ID in wb.sheetnames
    except Exception:  # noqa: BLE001
        return False


def lire_fiche_liaison(path: Path) -> tuple[int | None, dict[str, dict]]:
    """Lit un fichier de liaison rempli : renvoie (dossier_id, {code: réponse})."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)   # non read_only pour la robustesse
    dossier_id = None
    if FEUILLE_ID in wb.sheetnames:
        for row in wb[FEUILLE_ID].iter_rows(values_only=True):
            if row and str(row[0]).strip().lower() == "dossier_id" and len(row) > 1 and row[1] is not None:
                try:
                    dossier_id = int(row[1])
                except (TypeError, ValueError):
                    dossier_id = None
    reponses: dict[str, dict] = {}
    if FEUILLE_DEMANDES in wb.sheetnames:
        rows = list(wb[FEUILLE_DEMANDES].iter_rows(values_only=True))
        for r in rows[1:]:
            code = _cellule(r, 0)
            if not code:
                continue
            statut = _cellule(r, 3)       # « Statut (à mettre à jour) » — position déclarée
            reponse = _cellule(r, 4)      # « Réponse du candidat »
            reference = _cellule(r, 5)    # « Référence / pièce jointe »
            # preuve textuelle (réponse/référence) => alimente la ré-instruction Albert
            if reponse or reference:
                reponses[code] = {"mesure": statut, "precisions": reponse, "consolidation": reference}
    return dossier_id, reponses


def lire_statuts_liaison(path: Path) -> dict[str, str]:
    """Statut déclaré (colonne du menu déroulant) pour chaque exigence de la feuille."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    declares: dict[str, str] = {}
    if FEUILLE_DEMANDES in wb.sheetnames:
        for r in list(wb[FEUILLE_DEMANDES].iter_rows(values_only=True))[1:]:
            code = _cellule(r, 0)
            if code:
                declares[code] = _cellule(r, 3)
    return declares
