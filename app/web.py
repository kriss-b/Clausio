"""Interface web de Clausio (style DSFR — Suite numérique de l'État).

Pages : connexion, tableau de bord, nouvelle analyse (glisser-déposer + pré-remplissage),
détail d'un dossier (validation + rapport), administration (paramètres + édition dossiers),
mise à jour de dossier (réimport du fichier Excel de liaison). Les API /api/* servent
le JavaScript de ces pages. L'analyse tourne en tâche de fond avec suivi d'avancement.
"""
from __future__ import annotations

import json
import shutil
import threading
import uuid
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlmodel import Session, select

from . import analysis, auth, clausier_io, prefill, reporting, storage
from .config import VERSION
from .database import get_session
from .models import (Constat, Demande, Dossier, Document, Evenement, ExigenceRef, Parametres,
                     PhaseDocument, ReferentielVersion, Statut, StatutDossier, Utilisateur)
from .scoring import indicateur_global, indicateur_nis2, indicateur_rgpd

router = APIRouter()
BASE = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE / "templates"))
templates.env.globals["version"] = VERSION
STAGING = BASE / "staging"
UPLOADS = BASE / "uploads"
EXPORTS = BASE / "exports"

STATUTS_UI = ["couvert", "partiel", "absent", "non_applicable", "a_verifier"]
_BADGE_STATUT = {"couvert": "fr-badge--success", "partiel": "fr-badge--warning",
                 "absent": "fr-badge--error", "a_verifier": "fr-badge--warning",
                 "non_applicable": "fr-badge--info", "non_evalue": "fr-badge--grey"}
_BADGE_CRIT = {"bloquant": "fr-badge--error", "majeur": "fr-badge--warning", "mineur": "fr-badge--info"}
PROFIL_LABELS = {"socle": "Socle (hors DM connecté)",
                 "dispositif_medical": "Dispositif médical connecté (+ AFIB)"}
_RESOLU = {Statut.couvert, Statut.non_applicable}
# « à vérifier » (et non évalué) = état nul, en attente de qualification par le prestataire.
# Tout autre statut retenu est considéré comme TRAITÉ dans l'avancement.
_EN_ATTENTE = {Statut.a_verifier, Statut.non_evalue}


def session_dep():
    with get_session() as s:
        yield s


def _ctx(request: Request, **extra):
    nom = auth.utilisateur_courant(request)
    est_admin_global = False
    if nom:
        with get_session() as s:
            u = s.exec(select(Utilisateur).where(Utilisateur.username == nom)).first()
            est_admin_global = bool(u and u.role == "admin" and u.actif)
    base = {"utilisateur": nom, "est_admin_global": est_admin_global}
    base.update(extra)
    return base


def _analyser_en_fond(dossier_id: int, codes: set[str] | None = None):
    """Relance l'analyse dans un thread avec sa propre session (SQLite)."""
    def run():
        with get_session() as s:
            d = s.get(Dossier, dossier_id)
            if d:
                analysis.analyser(s, d, acteur="clausio", codes_cibles=codes)
    threading.Thread(target=run, daemon=True).start()


def _racine(s: Session, d: Dossier) -> Dossier:
    return s.get(Dossier, d.parent_id) if d.parent_id else d


def _visible(s: Session, request: Request, dossier: Dossier) -> bool:
    """Un dossier est visible par l'admin, son propriétaire (RSSI) ou son correspondant."""
    return auth.peut_voir(request, s, dossier)


def _dossier_visible(s: Session, request: Request, dossier_id: int) -> Dossier:
    d = s.get(Dossier, dossier_id)
    if not d:
        raise HTTPException(404, "Dossier inconnu.")
    if not _visible(s, request, d):
        raise HTTPException(403, "Vous n'avez pas accès à ce dossier.")
    return d


# ---------------------------------------------------------------- Auth
@router.get("/login", response_class=HTMLResponse)
def page_login(request: Request):
    return templates.TemplateResponse(request, "login.html", {})


@router.post("/login")
def faire_login(request: Request, utilisateur: str = Form(...), mot_de_passe: str = Form(...),
                s: Session = Depends(session_dep)):
    if auth.verifier_identifiants(utilisateur, mot_de_passe):
        u = s.exec(select(Utilisateur).where(Utilisateur.username == utilisateur)).first()
        if u and u.mfa_active and u.mfa_secret:
            request.session.pop("utilisateur", None)
            request.session["mfa_pending"] = utilisateur     # connexion en attente du code
            return templates.TemplateResponse(request, "mfa.html", {})
        request.session["utilisateur"] = utilisateur
        request.session.pop("mfa_pending", None)
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse(request, "login.html",
                                      {"erreur": "Identifiant ou mot de passe incorrect."})


@router.post("/login/mfa")
def faire_login_mfa(request: Request, code: str = Form(...), s: Session = Depends(session_dep)):
    nom = request.session.get("mfa_pending")
    if not nom:
        return RedirectResponse("/login", status_code=303)
    u = s.exec(select(Utilisateur).where(Utilisateur.username == nom)).first()
    if u and u.mfa_active and auth.mfa_verifier(u.mfa_secret, code):
        request.session["utilisateur"] = nom
        request.session.pop("mfa_pending", None)
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse(request, "mfa.html",
                                      {"erreur": "Code de vérification incorrect."})


@router.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


# ---------------------------------------------------------------- Tableau de bord
@router.get("/", response_class=HTMLResponse)
def dashboard(request: Request, s: Session = Depends(session_dep)):
    if (r := auth.exiger_connexion(request)):
        return r
    dossiers = s.exec(select(Dossier).where(Dossier.parent_id == None)  # noqa: E711
                      .order_by(Dossier.created_at.desc())).all()
    dossiers = [d for d in dossiers if _visible(s, request, d)]
    lignes = []
    for d in dossiers:
        nb_analyses = 1 + len(s.exec(select(Dossier).where(Dossier.parent_id == d.id)).all())
        constats = s.exec(select(Constat).where(Constat.dossier_id == d.id)).all()
        total = len(constats)
        valides = sum(1 for c in constats if c.statut_valide is not None)
        # « traitées » = état RETENU qualifié (tout sauf « à vérifier » / non évalué)
        resolues = sum(1 for c in constats
                       if (c.statut_valide if c.statut_valide is not None else c.statut_propose) not in _EN_ATTENTE)
        en_attente = total - resolues
        pct = round(100 * resolues / total) if total else 0
        if not d.analyse_termine:
            stade, badge = "analyse en cours", "fr-badge--info"
        elif d.statut == StatutDossier.clos:
            stade, badge = "clos", "fr-badge--success"
        elif total == 0:
            stade, badge = "à analyser", "fr-badge--grey"
        elif d.statut == StatutDossier.en_attente_complements and resolues < total:
            stade, badge = "en attente de compléments", "fr-badge--warning"
        elif valides == 0:
            stade, badge = "à instruire", "fr-badge--info"
        elif resolues < total:
            stade, badge = "en cours d'instruction", "fr-badge--new"
        else:
            stade, badge = "instruit", "fr-badge--success"
        lignes.append({"id": d.id, "nom": d.nom_affiche, "reference_marche": d.reference_marche,
                       "type_dispositif": d.type_dispositif,
                       "profil": PROFIL_LABELS.get(d.profil, d.profil),
                       "nb_analyses": nb_analyses,
                       "total": total, "resolues": resolues, "en_attente": en_attente,
                       "pct": pct, "stade": stade, "badge": badge})
    return templates.TemplateResponse(request, "dashboard.html",
        _ctx(request, dossiers=lignes, est_admin=auth.est_admin(request, s)))


@router.get("/nouvelle-analyse", response_class=HTMLResponse)
def page_nouvelle_analyse(request: Request):
    if (r := auth.exiger_connexion(request)):
        return r
    return templates.TemplateResponse(request, "nouvelle_analyse.html", _ctx(request))


@router.get("/mise-a-jour", response_class=HTMLResponse)
def page_mise_a_jour(request: Request):
    if (r := auth.exiger_connexion(request)):
        return r
    return templates.TemplateResponse(request, "mise_a_jour.html", _ctx(request))


@router.get("/dossiers/{dossier_id}/maj-resultat", response_class=HTMLResponse)
def page_maj_resultat(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if (r := auth.exiger_connexion(request)):
        return r
    dossier = _dossier_visible(s, request, dossier_id)
    return templates.TemplateResponse(request, "maj_resultat.html",
        _ctx(request, dossier=dossier))


@router.get("/dossiers/{dossier_id}", response_class=HTMLResponse)
def page_dossier(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if (r := auth.exiger_connexion(request)):
        return r
    dossier = _dossier_visible(s, request, dossier_id)

    # Famille d'analyses : dossier racine + analyses liées (onglets)
    racine_id = dossier.parent_id or dossier.id
    membres = [s.get(Dossier, racine_id)] + list(s.exec(
        select(Dossier).where(Dossier.parent_id == racine_id).order_by(Dossier.created_at)).all())
    refs_map = {r.id: r for r in s.exec(select(ReferentielVersion)).all()}
    onglets = []
    for m in membres:
        if not m:
            continue
        rv = refs_map.get(m.referentiel_version_id)
        onglets.append({"id": m.id, "actif": m.id == dossier.id,
                        "libelle": (rv.libelle.split("—")[0].strip() if rv else "Analyse"),
                        "profil": m.profil})
    rv_courant = refs_map.get(dossier.referentiel_version_id)
    ref_courant = rv_courant.libelle if rv_courant else ""

    # indicateur de conformité agrégé sur toutes les analyses de la candidature
    famille_ids = [m.id for m in membres if m]
    tous_constats = s.exec(select(Constat).where(Constat.dossier_id.in_(famille_ids))).all()
    conf_globale = indicateur_global(tous_constats)
    conf_globale["exigences"] = len(tous_constats)

    constats = s.exec(select(Constat).where(Constat.dossier_id == dossier_id)).all()
    libelles = {e.id: e.libelle for e in s.exec(
        select(ExigenceRef).where(ExigenceRef.referentiel_version_id == dossier.referentiel_version_id)).all()}
    groupes: dict[str, list] = {}
    for c in sorted(constats, key=lambda x: x.code_exigence):
        propose = c.statut_propose.value
        valide = c.statut_valide.value if c.statut_valide else None
        declare = c.statut_declare.value if c.statut_declare else None
        # présélection de la décision : déclaré par le candidat > décision existante > proposition
        defaut = valide or declare or propose
        groupes.setdefault(c.axe, []).append({
            "id": c.id, "code_exigence": c.code_exigence,
            "libelle": (libelles.get(c.exigence_ref_id, "") or "")[:220],
            "criticite_effective": c.criticite_effective.value,
            "statut_propose": propose,
            "statut_valide": valide,
            "statut_declare": declare,
            "valide": valide is not None,
            "statut_defaut": defaut,
            "badge_crit": _BADGE_CRIT.get(c.criticite_effective.value, "fr-badge--info"),
            "badge_prop": _BADGE_STATUT.get(propose, "fr-badge--grey"),
            "badge_decl": _BADGE_STATUT.get(declare, "fr-badge--grey"),
        })
    docs = s.exec(select(Document).where(Document.dossier_id == dossier_id)).all()
    documents = [{"id": d.id, "nom": d.nom, "type": d.type} for d in docs]
    return templates.TemplateResponse(request, "dossier.html",
        _ctx(request, dossier=dossier, groupes=groupes, statuts=STATUTS_UI,
             profil_label=PROFIL_LABELS.get(dossier.profil, dossier.profil),
             onglets=onglets, ref_courant=ref_courant, racine_id=racine_id,
             conf_globale=conf_globale, nb_analyses=len(famille_ids),
             est_admin=auth.est_admin(request, s), documents=documents))


@router.post("/api/dossiers/{dossier_id}/analyse-liee")
def creer_analyse_liee(dossier_id: int, request: Request,
                       referentiel_version_id: int = Form(...), profil: str = Form("socle"),
                       s: Session = Depends(session_dep)):
    """Crée une analyse d'un AUTRE référentiel, rattachée à la même candidature.
    Réutilise les mêmes documents ; l'analyse tourne en tâche de fond."""
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    base = _dossier_visible(s, request, dossier_id)
    if not s.get(ReferentielVersion, referentiel_version_id):
        raise HTTPException(404, "Référentiel inconnu.")
    racine_id = base.parent_id or base.id
    racine = s.get(Dossier, racine_id)

    enfant = Dossier(
        reference_marche=base.reference_marche, objet=base.objet,
        profil=profil, referentiel_version_id=referentiel_version_id,
        statut=StatutDossier.en_instruction, societe_ou_produit=base.societe_ou_produit,
        contact_nom=base.contact_nom, contact_email=base.contact_email, contact_tel=base.contact_tel,
        type_dispositif=base.type_dispositif, resume_ia=base.resume_ia,
        marche_nom=base.marche_nom, marche_email=base.marche_email, marche_tel=base.marche_tel,
        parent_id=racine_id, analyse_termine=False,
        owner_id=(racine.owner_id if racine else None),
        correspondant_id=(racine.correspondant_id if racine else None),
    )
    s.add(enfant)
    s.commit()
    s.refresh(enfant)

    # réutiliser les mêmes documents (mêmes fichiers sur disque) pour le corpus
    docs = s.exec(select(Document).where(Document.dossier_id == racine_id)).all()
    for d in docs:
        s.add(Document(dossier_id=enfant.id, nom=d.nom, type=d.type,
                       chemin=d.chemin, phase=d.phase))
    s.commit()

    _analyser_en_fond(enfant.id)
    return {"dossier_id": enfant.id}


@router.get("/dossiers/{dossier_id}/analyse-complementaire", response_class=HTMLResponse)
def page_analyse_complementaire(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if (r := auth.exiger_connexion(request)):
        return r
    dossier = _dossier_visible(s, request, dossier_id)
    racine_id = dossier.parent_id or dossier.id
    # membres déjà dans la famille (à exclure des rattachements possibles)
    famille = {racine_id} | {m.id for m in s.exec(
        select(Dossier).where(Dossier.parent_id == racine_id)).all()}
    refs = {r.id: r for r in s.exec(select(ReferentielVersion)).all()}
    # dossiers racines rattachables (autres candidatures/analyses indépendantes)
    rattachables = []
    for d in s.exec(select(Dossier).where(Dossier.parent_id == None)).all():  # noqa: E711
        if d.id in famille or not _visible(s, request, d):
            continue
        rv = refs.get(d.referentiel_version_id)
        rattachables.append({"id": d.id, "nom": d.nom_affiche, "reference": d.reference_marche,
                             "ref": (rv.libelle.split("—")[0].strip() if rv else "?"), "profil": d.profil})
    racine = s.get(Dossier, racine_id)
    return templates.TemplateResponse(request, "analyse_complementaire.html",
        _ctx(request, dossier=racine, rattachables=rattachables))


@router.post("/api/dossiers/{dossier_id}/rattacher")
def rattacher_analyse(dossier_id: int, request: Request, autre_id: int = Form(...),
                      s: Session = Depends(session_dep)):
    """Rattache un dossier existant (analyse indépendante) à la famille de dossier_id."""
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    base = _dossier_visible(s, request, dossier_id)
    autre = _dossier_visible(s, request, autre_id)
    racine_id = base.parent_id or base.id
    if autre.id == racine_id:
        raise HTTPException(400, "Un dossier ne peut pas être rattaché à lui-même.")
    # l'autre dossier et ses éventuelles analyses liées rejoignent la famille
    for enfant in s.exec(select(Dossier).where(Dossier.parent_id == autre.id)).all():
        enfant.parent_id = racine_id
        s.add(enfant)
    autre.parent_id = racine_id
    s.add(autre)
    s.commit()
    return {"dossier_id": autre.id, "racine_id": racine_id}


@router.get("/installation", response_class=HTMLResponse)
def page_installation(request: Request, s: Session = Depends(session_dep)):
    from . import settings
    if settings.installation_faite():
        return RedirectResponse("/", status_code=303)
    return templates.TemplateResponse(request, "installation.html", {"request": request})


class TestLLM(BaseModel):
    base_url: str
    api_key: str = ""


@router.post("/installation/test-llm")
def installation_test_llm(corps: TestLLM):
    from . import llm
    return llm.tester_connexion(corps.base_url, corps.api_key)


class UtilisateurInstall(BaseModel):
    username: str
    mot_de_passe: str = ""
    role: str = "utilisateur"
    nom: str = ""
    email: str = ""
    tel_fixe: str = ""
    tel_mobile: str = ""


class InstallationPayload(BaseModel):
    admin_username: str
    admin_mot_de_passe: str
    admin_nom: str = ""
    admin_email: str = ""
    admin_tel_fixe: str = ""
    admin_tel_mobile: str = ""
    llm_base_url: str = ""
    llm_api_key: str = ""
    llm_model: str = ""
    llm_embed_model: str = ""
    auth_mode: str = "local"
    ldap_host: str = ""
    ldap_port: int = 389
    ldap_use_tls: bool = True
    ldap_base_dn: str = ""
    ldap_bind_template: str = ""
    utilisateurs: list[UtilisateurInstall] = []


@router.post("/installation")
def installation_finaliser(corps: InstallationPayload, s: Session = Depends(session_dep)):
    from . import settings, llm
    cfg = settings.charger_config(s)
    if cfg.installation_faite:
        raise HTTPException(409, "Installation déjà effectuée.")
    if not corps.admin_username.strip() or not corps.admin_mot_de_passe:
        raise HTTPException(422, "Identifiant et mot de passe administrateur requis.")

    # 1) compte admin
    salt, h = auth.hacher(corps.admin_mot_de_passe)
    s.add(Utilisateur(username=corps.admin_username.strip(), nom=corps.admin_nom or "Administrateur",
                      email=corps.admin_email, tel_fixe=corps.admin_tel_fixe, tel_mobile=corps.admin_tel_mobile,
                      role="admin", salt=salt, mot_de_passe_hash=h, actif=True))

    # 2) comptes supplémentaires
    for u in corps.utilisateurs:
        if not u.username.strip():
            continue
        if s.exec(select(Utilisateur).where(Utilisateur.username == u.username.strip())).first():
            continue
        us, uh = auth.hacher(u.mot_de_passe) if u.mot_de_passe else ("", "")
        s.add(Utilisateur(username=u.username.strip(), nom=u.nom, email=u.email,
                          tel_fixe=u.tel_fixe, tel_mobile=u.tel_mobile,
                          role=("admin" if u.role == "admin" else "utilisateur"),
                          salt=us, mot_de_passe_hash=uh, actif=True))

    # 3) LLM + authentification
    cfg.llm_base_url = llm._normaliser_base(corps.llm_base_url)
    cfg.llm_api_key = corps.llm_api_key.strip()
    cfg.llm_model = corps.llm_model.strip()
    cfg.llm_embed_model = corps.llm_embed_model.strip()
    cfg.auth_mode = "ldap" if corps.auth_mode == "ldap" else "local"
    cfg.ldap_host = corps.ldap_host.strip()
    cfg.ldap_port = corps.ldap_port or 389
    cfg.ldap_use_tls = bool(corps.ldap_use_tls)
    cfg.ldap_base_dn = corps.ldap_base_dn.strip()
    cfg.ldap_bind_template = corps.ldap_bind_template.strip()
    cfg.installation_faite = True
    s.add(cfg)
    s.commit()
    llm.reset_cache()
    return {"ok": True}


# ---------------------------------------------------------------- Administration
@router.get("/administration", response_class=HTMLResponse)
def page_admin(request: Request, s: Session = Depends(session_dep)):
    if (r := auth.exiger_connexion(request)):
        return r
    moi = auth.utilisateur_obj(request, s)
    est_admin = bool(moi and moi.role == "admin")
    params = s.get(Parametres, 1) or Parametres(id=1)
    utilisateurs = s.exec(select(Utilisateur).order_by(Utilisateur.username)).all() if est_admin else []
    from . import settings as _st
    cfg = _st.charger_config(s)
    ia = {"base_url": cfg.llm_base_url, "model": cfg.llm_model, "embed_model": cfg.llm_embed_model,
          "temperature": cfg.llm_temperature or 0.0, "a_cle": bool(cfg.llm_api_key),
          "auth_mode": cfg.auth_mode} if est_admin else {}
    ref_repo_url = (cfg.ref_repo_url or "https://github.com/nocomp/Clausio/tree/main/referentiels") if est_admin else ""
    smtp = {"actif": cfg.smtp_actif, "host": cfg.smtp_host, "port": cfg.smtp_port or 587,
            "user": cfg.smtp_user, "from": cfg.smtp_from, "securite": cfg.smtp_securite or "starttls",
            "a_mdp": bool(cfg.smtp_password), "base_url": cfg.app_base_url} if est_admin else {}
    dossiers = [d for d in s.exec(select(Dossier).order_by(Dossier.created_at.desc())).all()
                if _visible(s, request, d)]
    return templates.TemplateResponse(request, "administration.html",
        _ctx(request, params=params, dossiers=dossiers, utilisateurs=utilisateurs,
             est_admin=est_admin, moi=moi, ia=ia, ref_repo_url=ref_repo_url, smtp=smtp))


class ConfigSMTP(BaseModel):
    smtp_actif: bool = False
    smtp_host: str = ""
    smtp_port: int = 587
    smtp_user: str = ""
    smtp_password: str = ""
    smtp_from: str = ""
    smtp_securite: str = "starttls"
    app_base_url: str = ""


@router.post("/administration/smtp")
def enregistrer_smtp(corps: ConfigSMTP, request: Request, s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403)
    from . import settings as _st
    cfg = _st.charger_config(s)
    cfg.smtp_actif = bool(corps.smtp_actif)
    cfg.smtp_host = corps.smtp_host.strip()
    cfg.smtp_port = corps.smtp_port or 587
    cfg.smtp_user = corps.smtp_user.strip()
    if corps.smtp_password.strip():         # ne pas écraser si laissé vide
        cfg.smtp_password = corps.smtp_password
    cfg.smtp_from = corps.smtp_from.strip()
    cfg.smtp_securite = corps.smtp_securite if corps.smtp_securite in ("starttls", "ssl", "aucune") else "starttls"
    cfg.app_base_url = corps.app_base_url.strip()
    s.add(cfg)
    s.commit()
    return {"ok": True}


class TestSMTP(BaseModel):
    destinataire: str = ""


@router.post("/administration/smtp/test")
def tester_smtp(corps: TestSMTP, request: Request, s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403)
    from . import notifier
    dest = corps.destinataire.strip()
    if not dest:
        moi = auth.utilisateur_obj(request, s)
        dest = moi.email if moi else ""
    if not dest:
        return {"ok": False, "erreur": "Renseignez un destinataire (ou l'adresse de votre compte)."}
    return notifier.tester(dest)


class MajReferentiels(BaseModel):
    url: str = ""


@router.post("/administration/referentiels/maj")
def maj_referentiels(corps: MajReferentiels, request: Request, s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403)
    from . import referentiel, settings as _st
    cfg = _st.charger_config(s)
    url = corps.url.strip() or cfg.ref_repo_url or "https://github.com/nocomp/Clausio/tree/main/referentiels"
    res = referentiel.maj_depuis_github(s, url)
    if res.get("ok"):
        cfg.ref_repo_url = url
        s.add(cfg)
        s.commit()
    return res


class CodeMFA(BaseModel):
    code: str = ""


@router.post("/administration/mfa/init")
def mfa_init(request: Request, s: Session = Depends(session_dep)):
    """Génère un secret TOTP pour le compte courant et renvoie le QR (SVG) à scanner."""
    u = auth.utilisateur_obj(request, s)
    if not u:
        raise HTTPException(401)
    import pyotp
    import qrcode
    secret = pyotp.random_base32()
    u.mfa_secret = secret
    u.mfa_active = False                    # pas actif tant que le code n'est pas confirmé
    s.add(u)
    s.commit()
    uri = pyotp.TOTP(secret).provisioning_uri(name=u.username, issuer_name="Clausio")
    # SVG construit à la main depuis la matrice (rendu fiable, sans prologue XML ni préfixe) :
    qr = qrcode.QRCode(border=2, error_correction=qrcode.constants.ERROR_CORRECT_M)
    qr.add_data(uri)
    qr.make(fit=True)
    matrice = qr.get_matrix()
    n = len(matrice)
    modules = "".join(
        f'<rect x="{x}" y="{y}" width="1" height="1"/>'
        for y, ligne in enumerate(matrice) for x, plein in enumerate(ligne) if plein
    )
    svg = (f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {n} {n}" '
           f'width="100%" shape-rendering="crispEdges" role="img" '
           f'aria-label="QR code d\'activation">'
           f'<rect width="{n}" height="{n}" fill="#ffffff"/>'
           f'<g fill="#000091">{modules}</g></svg>')
    return {"secret": secret, "svg": svg}


@router.post("/administration/mfa/activer")
def mfa_activer(corps: CodeMFA, request: Request, s: Session = Depends(session_dep)):
    u = auth.utilisateur_obj(request, s)
    if not u:
        raise HTTPException(401)
    if auth.mfa_verifier(u.mfa_secret, corps.code):
        u.mfa_active = True
        s.add(u)
        s.commit()
        return {"ok": True}
    return {"ok": False, "erreur": "Code invalide — vérifiez l'heure de votre téléphone et réessayez."}


@router.post("/administration/mfa/desactiver")
def mfa_desactiver(request: Request, s: Session = Depends(session_dep)):
    u = auth.utilisateur_obj(request, s)
    if not u:
        raise HTTPException(401)
    u.mfa_active = False
    u.mfa_secret = ""
    s.add(u)
    s.commit()
    return {"ok": True}


class ConfigIA(BaseModel):
    llm_base_url: str = ""
    llm_api_key: str = ""
    llm_model: str = ""
    llm_embed_model: str = ""
    llm_temperature: float = 0.0


@router.post("/administration/ia")
def enregistrer_ia(corps: ConfigIA, request: Request, s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403)
    from . import settings as _st, llm
    cfg = _st.charger_config(s)
    cfg.llm_base_url = llm._normaliser_base(corps.llm_base_url)
    if corps.llm_api_key.strip():           # ne pas écraser la clé si le champ est laissé vide
        cfg.llm_api_key = corps.llm_api_key.strip()
    cfg.llm_model = corps.llm_model.strip()
    cfg.llm_embed_model = corps.llm_embed_model.strip()
    try:
        cfg.llm_temperature = max(0.0, min(2.0, float(corps.llm_temperature)))
    except (TypeError, ValueError):
        cfg.llm_temperature = 0.0
    s.add(cfg)
    s.commit()
    llm.reset_cache()
    return {"ok": True}


@router.post("/administration/ia/test")
def tester_ia(corps: TestLLM, request: Request, s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403)
    from . import llm, settings as _st
    if not corps.base_url.strip():          # test avec la config enregistrée si champ vide
        cfg = _st.charger_config(s)
        return llm.tester_connexion(cfg.llm_base_url, cfg.llm_api_key)
    return llm.tester_connexion(corps.base_url, corps.api_key)


@router.post("/administration/utilisateurs")
def creer_utilisateur(request: Request, username: str = Form(...), nom: str = Form(""),
                      email: str = Form(""), role: str = Form("utilisateur"),
                      mot_de_passe: str = Form(...), mot_de_passe2: str = Form(""),
                      tel_fixe: str = Form(""), tel_mobile: str = Form(""),
                      s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403)
    username = username.strip()
    if not username or not mot_de_passe:
        raise HTTPException(422, "Identifiant et mot de passe requis.")
    if mot_de_passe2 and mot_de_passe != mot_de_passe2:
        raise HTTPException(422, "Les deux mots de passe ne correspondent pas.")
    if s.exec(select(Utilisateur).where(Utilisateur.username == username)).first():
        raise HTTPException(409, "Cet identifiant existe déjà.")
    salt, h = auth.hacher(mot_de_passe)
    s.add(Utilisateur(username=username, nom=nom, email=email,
                      tel_fixe=tel_fixe.strip(), tel_mobile=tel_mobile.strip(),
                      role=("admin" if role == "admin" else "utilisateur"),
                      salt=salt, mot_de_passe_hash=h, actif=True))
    s.commit()
    return RedirectResponse("/administration", status_code=303)


@router.post("/administration/utilisateurs/{uid}/reset")
def reset_mdp(uid: int, request: Request, nouveau: str = Form(...),
              s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403)
    u = s.get(Utilisateur, uid)
    if not u:
        raise HTTPException(404)
    u.salt, u.mot_de_passe_hash = auth.hacher(nouveau)
    s.add(u)
    s.commit()
    return RedirectResponse("/administration", status_code=303)


@router.post("/administration/utilisateurs/{uid}/actif")
def toggle_actif(uid: int, request: Request, s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403)
    u = s.get(Utilisateur, uid)
    moi = auth.utilisateur_obj(request, s)
    if not u:
        raise HTTPException(404)
    if moi and u.id == moi.id:
        raise HTTPException(400, "Vous ne pouvez pas désactiver votre propre compte.")
    u.actif = not u.actif
    s.add(u)
    s.commit()
    return RedirectResponse("/administration", status_code=303)


@router.post("/administration/mon-mot-de-passe")
def changer_mon_mdp(request: Request, ancien: str = Form(...), nouveau: str = Form(...),
                    s: Session = Depends(session_dep)):
    moi = auth.utilisateur_obj(request, s)
    if not moi:
        raise HTTPException(401)
    if not auth.verifier_mdp(ancien, moi.salt, moi.mot_de_passe_hash):
        raise HTTPException(403, "Ancien mot de passe incorrect.")
    if not nouveau:
        raise HTTPException(422, "Nouveau mot de passe requis.")
    moi.salt, moi.mot_de_passe_hash = auth.hacher(nouveau)
    s.add(moi)
    s.commit()
    return RedirectResponse("/administration", status_code=303)


@router.get("/dossiers/{dossier_id}/diagnostic.xlsx")
def export_diagnostic(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    """Export d'audit (admin) : pour chaque exigence, statut proposé par Albert, confiance,
    justification, et les passages qui lui ont été soumis. Sert à comprendre les 'à vérifier'."""
    if not auth.est_admin(request, s):
        raise HTTPException(403, "Réservé à l'administrateur.")
    d = s.get(Dossier, dossier_id)
    if not d:
        raise HTTPException(404)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from .models import Preuve
    libelles = {e.id: e.libelle for e in s.exec(
        select(ExigenceRef).where(ExigenceRef.referentiel_version_id == d.referentiel_version_id)).all()}
    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnostic"
    entetes = ["Code", "Exigence", "Criticité", "Statut proposé", "Confiance",
               "Justification IA", "Passages soumis à Albert (source → extrait)"]
    ws.append(entetes)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = c.fill.copy(patternType="solid", fgColor="000091")
    constats = s.exec(select(Constat).where(Constat.dossier_id == dossier_id)
                      .order_by(Constat.code_exigence)).all()
    for c in constats:
        preuves = s.exec(select(Preuve).where(Preuve.constat_id == c.id)).all()
        passages = "\n---\n".join(
            f"[{p.document_nom} p.{p.page or '?'}] {p.extrait}" for p in preuves) or "(aucun passage retrouvé)"
        ws.append([c.code_exigence, libelles.get(c.exigence_ref_id, ""),
                   c.criticite_effective.value,
                   c.statut_propose.value if c.statut_propose else "",
                   round(c.confiance_ia or 0, 2), c.justification_ia or "", passages])
    largeurs = [12, 50, 12, 16, 10, 60, 90]
    for i, w in enumerate(largeurs, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    dest = storage.dir_exports(s, dossier_id)
    chemin = dest / f"diagnostic_dossier_{dossier_id}.xlsx"
    wb.save(chemin)
    return FileResponse(chemin, filename=f"Clausio_diagnostic_{dossier_id}.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.get("/api/utilisateurs")
def liste_utilisateurs(request: Request, s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    us = s.exec(select(Utilisateur).where(Utilisateur.actif == True)  # noqa: E712
                .order_by(Utilisateur.username)).all()
    return [{"id": u.id, "username": u.username, "nom": u.nom} for u in us]


def _effacer_dossier_db(s: Session, d: Dossier) -> None:
    from .models import Preuve
    for c in s.exec(select(Constat).where(Constat.dossier_id == d.id)).all():
        for pr in s.exec(select(Preuve).where(Preuve.constat_id == c.id)).all():
            s.delete(pr)
        s.delete(c)
    for dm in s.exec(select(Demande).where(Demande.dossier_id == d.id)).all():
        s.delete(dm)
    for ev in s.exec(select(Evenement).where(Evenement.dossier_id == d.id)).all():
        s.delete(ev)
    for doc in s.exec(select(Document).where(Document.dossier_id == d.id)).all():
        s.delete(doc)
    shutil.rmtree(storage.dir_exports(s, d.id), ignore_errors=True)
    s.delete(d)


def _effacer_famille(s: Session, root_id: int) -> None:
    for e in s.exec(select(Dossier).where(Dossier.parent_id == root_id)).all():
        _effacer_dossier_db(s, e)
    root = s.get(Dossier, root_id)
    if root:
        _effacer_dossier_db(s, root)
    shutil.rmtree(storage.dir_uploads(s, root_id), ignore_errors=True)
    s.commit()


class SupprimerDossiers(BaseModel):
    ids: list[int]


@router.post("/api/dossiers/supprimer")
def supprimer_dossiers(corps: SupprimerDossiers, request: Request, s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403, "Réservé à l'administrateur.")
    n = 0
    for did in corps.ids:
        d = s.get(Dossier, did)
        if not d:
            continue
        racine_id = d.parent_id or d.id
        _effacer_famille(s, racine_id)
        n += 1
    return {"supprimes": n}


@router.post("/api/dossiers/{dossier_id}/supprimer-analyse")
def supprimer_analyse(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403, "Réservé à l'administrateur.")
    d = s.get(Dossier, dossier_id)
    if not d:
        raise HTTPException(404)
    if d.parent_id is None:            # racine -> on efface toute la candidature
        _effacer_famille(s, d.id)
        return {"supprime": "famille"}
    _effacer_dossier_db(s, d)          # analyse liée -> on n'efface pas les fichiers partagés
    s.commit()
    return {"supprime": "analyse", "racine_id": d.parent_id}


@router.post("/api/dossiers/{dossier_id}/documents/{doc_id}/supprimer")
def supprimer_document(dossier_id: int, doc_id: int, request: Request, s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403, "Réservé à l'administrateur.")
    doc = s.get(Document, doc_id)
    if not doc or doc.dossier_id != dossier_id:
        raise HTTPException(404)
    chemin = doc.chemin
    s.delete(doc)
    s.commit()
    autres = s.exec(select(Document).where(Document.chemin == chemin)).all()
    if not autres:
        Path(chemin).unlink(missing_ok=True)
    return {"supprime": doc_id}


@router.post("/administration/parametres")
async def enregistrer_parametres(request: Request, etablissement: str = Form(""), rssi_nom: str = Form(""),
                                 rssi_email: str = Form(""), rssi_tel: str = Form(""),
                                 logo: UploadFile = File(None),
                                 s: Session = Depends(session_dep)):
    if not auth.est_admin(request, s):
        raise HTTPException(403)
    params = s.get(Parametres, 1) or Parametres(id=1)
    params.etablissement, params.rssi_nom = etablissement, rssi_nom
    params.rssi_email, params.rssi_tel = rssi_email, rssi_tel
    if logo is not None and logo.filename:
        ext = Path(logo.filename).suffix.lower()
        if ext in (".png", ".jpg", ".jpeg", ".gif"):
            dossier_logo = EXPORTS / "_params"
            dossier_logo.mkdir(parents=True, exist_ok=True)
            cible = dossier_logo / f"logo{ext}"
            cible.write_bytes(await logo.read())
            params.logo_path = str(cible.relative_to(BASE))
    s.add(params)
    s.commit()
    return RedirectResponse("/administration", status_code=303)


@router.get("/administration/logo")
def logo_etablissement(request: Request, s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    params = s.get(Parametres, 1)
    if not params or not params.logo_path:
        raise HTTPException(404)
    chemin = BASE / params.logo_path
    if not chemin.exists():
        raise HTTPException(404)
    return FileResponse(chemin)


@router.get("/administration/dossier/{dossier_id}", response_class=HTMLResponse)
def page_edit_dossier(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if (r := auth.exiger_connexion(request)):
        return r
    dossier = _dossier_visible(s, request, dossier_id)
    racine = _racine(s, dossier)
    utilisateurs = s.exec(select(Utilisateur).where(Utilisateur.actif == True).order_by(Utilisateur.username)).all()  # noqa: E712
    return templates.TemplateResponse(request, "edit_dossier.html",
        _ctx(request, dossier=dossier, utilisateurs=utilisateurs, correspondant_id=racine.correspondant_id))


@router.post("/administration/dossier/{dossier_id}")
def enregistrer_dossier(dossier_id: int, request: Request,
                        societe_ou_produit: str = Form(""), objet: str = Form(""),
                        reference_marche: str = Form(""), type_dispositif: str = Form(""),
                        contact_nom: str = Form(""), contact_email: str = Form(""), contact_tel: str = Form(""),
                        marche_nom: str = Form(""), marche_email: str = Form(""), marche_tel: str = Form(""),
                        correspondant_id: str = Form(""),
                        s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    d = _dossier_visible(s, request, dossier_id)
    racine = _racine(s, d)
    racine.correspondant_id = int(correspondant_id) if correspondant_id.strip().isdigit() else None
    s.add(racine)
    d.societe_ou_produit, d.objet, d.reference_marche = societe_ou_produit, objet, reference_marche
    d.type_dispositif = type_dispositif
    d.contact_nom, d.contact_email, d.contact_tel = contact_nom, contact_email, contact_tel
    d.marche_nom, d.marche_email, d.marche_tel = marche_nom, marche_email, marche_tel
    s.add(d)
    s.commit()
    return RedirectResponse("/administration", status_code=303)


# ---------------------------------------------------------------- API : nouvelle analyse
@router.post("/api/pre-analyse")
async def api_pre_analyse(request: Request, fichiers: list[UploadFile] = File(...)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401, "Non authentifié.")
    staging_id = uuid.uuid4().hex
    dest = STAGING / staging_id
    dest.mkdir(parents=True, exist_ok=True)
    chemins = []
    manifeste = {}
    for f in fichiers:
        data = await f.read()
        if not storage.taille_ok(data):
            shutil.rmtree(dest, ignore_errors=True)
            raise HTTPException(413, f"Fichier trop volumineux (max {storage.MAX_UPLOAD_MO} Mo).")
        stocke = storage.nom_stocke(f.filename)      # nom aléatoire non prévisible
        p = dest / stocke
        p.write_bytes(data)
        manifeste[stocke] = storage.nom_affichage(f.filename)
        chemins.append(p)
    (dest / "_manifest.json").write_text(json.dumps(manifeste), encoding="utf-8")
    return {"staging_id": staging_id, "prefill": prefill.pre_remplir(chemins)}


class CreerDossier(BaseModel):
    staging_id: str
    societe_ou_produit: str = ""
    objet: str = ""
    contact_nom: str = ""
    contact_email: str = ""
    contact_tel: str = ""
    type_dispositif: str = ""
    profil: str = "socle"
    referentiel_version_id: int
    reference_marche: str = "à préciser"
    resume_ia: str = ""
    marche_nom: str = ""
    marche_email: str = ""
    marche_tel: str = ""
    correspondant_id: Optional[int] = None


@router.post("/api/dossiers/creer")
def api_creer_dossier(corps: CreerDossier, request: Request, s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401, "Non authentifié.")
    if not s.get(ReferentielVersion, corps.referentiel_version_id):
        raise HTTPException(404, "Référentiel inconnu.")
    moi = auth.utilisateur_obj(request, s)

    dossier = Dossier(
        reference_marche=corps.reference_marche,
        objet=corps.objet or corps.societe_ou_produit or "Candidature",
        profil=corps.profil, referentiel_version_id=corps.referentiel_version_id,
        statut=StatutDossier.en_instruction, societe_ou_produit=corps.societe_ou_produit,
        contact_nom=corps.contact_nom, contact_email=corps.contact_email, contact_tel=corps.contact_tel,
        type_dispositif=corps.type_dispositif, resume_ia=corps.resume_ia,
        marche_nom=corps.marche_nom, marche_email=corps.marche_email, marche_tel=corps.marche_tel,
        analyse_termine=False,
        owner_id=(moi.id if moi else None), correspondant_id=corps.correspondant_id,
    )
    s.add(dossier)
    s.commit()
    s.refresh(dossier)

    src = STAGING / corps.staging_id
    dest = storage.dir_uploads(s, dossier.id)
    manifeste = {}
    mf = src / "_manifest.json"
    if mf.exists():
        try:
            manifeste = json.loads(mf.read_text(encoding="utf-8"))
        except Exception:  # noqa: BLE001
            manifeste = {}
    if src.exists():
        for f in src.iterdir():
            if f.name == "_manifest.json":
                continue
            cible = dest / f.name                 # nom déjà aléatoire (staging)
            shutil.move(str(f), str(cible))
            affichage = manifeste.get(f.name, f.name)
            s.add(Document(dossier_id=dossier.id, nom=affichage,
                           type=(affichage.rsplit(".", 1)[-1] if "." in affichage else ""),
                           chemin=str(cible), phase=PhaseDocument.initiale))
        s.commit()
        shutil.rmtree(src, ignore_errors=True)

    _analyser_en_fond(dossier.id)
    return {"dossier_id": dossier.id}


@router.get("/api/dossiers/{dossier_id}/progression")
def progression(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    d = _dossier_visible(s, request, dossier_id)
    total = d.analyse_total or 0
    faites = d.analyse_faites or 0
    return {"total": total, "faites": faites, "termine": bool(d.analyse_termine),
            "pct": round(100 * faites / total) if total else (100 if d.analyse_termine else 0)}


# ---------------------------------------------------------------- API : validation
@router.post("/api/dossiers/{dossier_id}/tout-valider")
def tout_valider(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    """Applique à chaque exigence le statut proposé comme décision RSSI."""
    utilisateur = auth.utilisateur_courant(request)
    if not utilisateur:
        raise HTTPException(401)
    _dossier_visible(s, request, dossier_id)
    from datetime import datetime, timezone
    constats = s.exec(select(Constat).where(Constat.dossier_id == dossier_id)).all()
    for c in constats:
        c.statut_valide = c.statut_propose
        c.valide_par = utilisateur
        c.valide_at = datetime.now(timezone.utc)
        s.add(c)
    s.commit()
    analysis._regenerer_demandes(s, s.get(Dossier, dossier_id))
    return {"valides": len(constats)}


# ---------------------------------------------------------------- API : rapport (PDF + Excel)
@router.post("/api/dossiers/{dossier_id}/rapport-fichiers")
def rapport_fichiers(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    dossier = _dossier_visible(s, request, dossier_id)
    from .models import Demande, Evenement
    livrables = reporting.generer_livrables(s, dossier)
    demandes = s.exec(select(Demande).where(Demande.dossier_id == dossier_id).order_by(Demande.numero)).all()
    livrables["demandes_complements"] = [
        {"numero": d.numero, "code": d.code_exigence, "libelle": d.libelle, "texte": d.texte}
        for d in demandes]
    livrables["pdf_url"] = f"/dossiers/{dossier_id}/rapport.pdf"
    livrables["excel_url"] = f"/dossiers/{dossier_id}/liaison.xlsx"
    livrables["zip_url"] = f"/dossiers/{dossier_id}/rapport.zip"
    livrables["avertissement"] = ("Pré-instruction outillée validée par le RSSI. "
                                  "La décision d'attribution demeure humaine et motivée.")
    # s'il reste des demandes ouvertes, le dossier passe « en attente de compléments »
    if demandes and dossier.statut != StatutDossier.clos:
        dossier.statut = StatutDossier.en_attente_complements
        s.add(dossier)
    s.add(Evenement(dossier_id=dossier_id, type="rapport",
                    acteur=auth.utilisateur_courant(request),
                    details={"conclusion": livrables["conclusion"]}))
    s.commit()
    return livrables


@router.post("/api/dossiers/{dossier_id}/cloturer")
def cloturer(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    d = _dossier_visible(s, request, dossier_id)
    d.statut = StatutDossier.clos
    s.add(d)
    from .models import Evenement
    s.add(Evenement(dossier_id=dossier_id, type="cloture", acteur=auth.utilisateur_courant(request)))
    s.commit()
    return {"statut": "clos"}


@router.get("/dossiers/{dossier_id}/rapport.zip")
def telecharger_zip(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    d = _dossier_visible(s, request, dossier_id)
    chemin = storage.dir_exports(s, dossier_id) / f"clausio_dossier_{dossier_id}.zip"
    if not chemin.exists() or not d:
        raise HTTPException(404, "Archive non générée.")
    return FileResponse(chemin, media_type="application/zip",
                        filename=reporting.noms_fichiers(d)["zip"])


@router.get("/dossiers/{dossier_id}/rapport.pdf")
def telecharger_pdf(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    d = _dossier_visible(s, request, dossier_id)
    chemin = storage.dir_exports(s, dossier_id) / f"rapport_dossier_{dossier_id}.pdf"
    if not chemin.exists() or not d:
        raise HTTPException(404, "Rapport non généré.")
    return FileResponse(chemin, media_type="application/pdf",
                        filename=reporting.noms_fichiers(d)["pdf"])


@router.get("/dossiers/{dossier_id}/liaison.xlsx")
def telecharger_liaison(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    d = _dossier_visible(s, request, dossier_id)
    chemin = storage.dir_exports(s, dossier_id) / f"liaison_dossier_{dossier_id}.xlsx"
    if not chemin.exists() or not d:
        raise HTTPException(404, "Fichier de liaison non généré.")
    return FileResponse(chemin, filename=reporting.noms_fichiers(d)["excel"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------- API : mise à jour (reprise)
@router.post("/api/dossiers/maj")
async def api_mise_a_jour(request: Request, fichier: UploadFile = File(...),
                          s: Session = Depends(session_dep)):
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    tmp = STAGING / f"maj_{uuid.uuid4().hex}.xlsx"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_bytes(await fichier.read())

    try:
        reconnu = clausier_io.est_fiche_liaison(tmp)
        dossier_id, reponses = clausier_io.lire_fiche_liaison(tmp) if reconnu else (None, {})
    except Exception:  # noqa: BLE001
        reconnu, dossier_id, reponses = False, None, {}
    if not reconnu:
        tmp.unlink(missing_ok=True)
        raise HTTPException(422, "Ce fichier n'est pas un fichier de liaison Clausio, ou il est illisible.")
    dossier = s.get(Dossier, dossier_id) if dossier_id else None
    if not dossier:
        tmp.unlink(missing_ok=True)
        raise HTTPException(404, "Dossier introuvable pour ce fichier de liaison.")
    if not _visible(s, request, dossier):
        tmp.unlink(missing_ok=True)
        raise HTTPException(403, "Vous n'avez pas accès à ce dossier.")

    # --- Instantané AVANT réanalyse (pour la vue des changements) ---
    def _retenu(c):
        return (c.statut_valide if c.statut_valide is not None else c.statut_propose).value

    constats_avant = s.exec(select(Constat).where(Constat.dossier_id == dossier.id)).all()
    avant = {c.code_exigence: _retenu(c) for c in constats_avant}
    indic_avant = {
        "global": indicateur_global(constats_avant),
        "nis2": indicateur_nis2(constats_avant),
        "rgpd": indicateur_rgpd(constats_avant),
        "levees": sum(1 for c in constats_avant
                      if (c.statut_valide if c.statut_valide is not None else c.statut_propose) not in _EN_ATTENTE),
        "total": len(constats_avant),
    }
    reponses_texte = {code: (r.get("precisions", "") or "") for code, r in reponses.items()}

    # statuts déclarés par le candidat (menu déroulant) — comparés à l'état courant
    try:
        declares_bruts = clausier_io.lire_statuts_liaison(tmp)
    except Exception:  # noqa: BLE001
        declares_bruts = {}
    courant = {c.code_exigence: _retenu(c) for c in constats_avant}
    declares = {}
    for code, decl in declares_bruts.items():
        decl = (decl or "").strip()
        if decl in STATUTS_UI and code in courant and decl != courant[code]:
            declares[code] = decl

    dest = storage.dir_uploads(s, dossier.id)
    cible = dest / storage.nom_stocke(fichier.filename or "liaison.xlsx")
    shutil.move(str(tmp), str(cible))
    s.add(Document(dossier_id=dossier.id, nom=storage.nom_affichage(fichier.filename or "liaison.xlsx"),
                   type="xlsx", chemin=str(cible), phase=PhaseDocument.complement))

    # Pour chaque exigence touchée : on ré-ouvre la décision RSSI et on enregistre la
    # position déclarée par le candidat. Cette position PRÉ-REMPLIT la décision RSSI
    # (menu du dossier) sans être appliquée d'office. Seule une réponse TEXTUELLE
    # déclenche une ré-instruction par Albert (nouvelle preuve à qualifier).
    codes_texte = {code for code, t in reponses_texte.items() if t}
    touched = sorted(set(reponses_texte) | set(declares))
    for code in touched:
        c = s.exec(select(Constat).where(Constat.dossier_id == dossier.id,
                                         Constat.code_exigence == code)).first()
        if not c:
            continue
        if code in declares:
            try:
                c.statut_declare = Statut(declares[code])
            except ValueError:
                pass
        # ré-ouvrir (la décision RSSI sera reprise, pré-remplie sur le déclaré)
        c.statut_valide = None
        c.valide_par = None
        c.valide_at = None
        s.add(c)

    dossier.analyse_termine = not bool(codes_texte)   # analyse de fond seulement si réponses
    s.add(dossier)
    s.add(Evenement(dossier_id=dossier.id, type="mise_a_jour",
                    acteur=auth.utilisateur_courant(request),
                    details={"avant": avant, "indic_avant": indic_avant,
                             "reponses": reponses_texte, "declares": declares,
                             "touched": touched}))
    s.commit()

    if codes_texte:
        _analyser_en_fond(dossier.id, codes=codes_texte)

    # Notification par courriel aux personnes du dossier (best-effort, tâche de fond)
    try:
        from . import notifier
        chgs = [{"code": code,
                 "avant": avant.get(code, "?"),
                 "apres": declares.get(code) or ("à réévaluer" if code in codes_texte else "?")}
                for code in touched]
        notifier.notifier_maj_dossier(dossier.id, chgs, auteur=auth.utilisateur_courant(request) or "")
    except Exception:  # noqa: BLE001
        pass

    return {"dossier_id": dossier.id, "nom": dossier.nom_affiche,
            "modifications": len(touched)}


@router.get("/api/dossiers/{dossier_id}/maj-diff")
def maj_diff(dossier_id: int, request: Request, s: Session = Depends(session_dep)):
    """Compare l'état avant/après la dernière mise à jour, pondéré par la criticité."""
    if not auth.utilisateur_courant(request):
        raise HTTPException(401)
    dossier = _dossier_visible(s, request, dossier_id)
    ev = s.exec(select(Evenement).where(Evenement.dossier_id == dossier_id,
                Evenement.type == "mise_a_jour").order_by(Evenement.at.desc())).first()
    if not ev:
        raise HTTPException(404, "Aucune mise à jour pour ce dossier.")
    avant = ev.details.get("avant", {})
    reponses = ev.details.get("reponses", {})
    declares = ev.details.get("declares", {})
    touched = ev.details.get("touched", sorted(set(reponses) | set(declares)))
    indic_avant = ev.details.get("indic_avant", {})

    constats = {c.code_exigence: c for c in
                s.exec(select(Constat).where(Constat.dossier_id == dossier_id)).all()}
    libelles = {e.id: e.libelle for e in s.exec(
        select(ExigenceRef).where(ExigenceRef.referentiel_version_id == dossier.referentiel_version_id)).all()}
    rang = {"bloquant": 0, "majeur": 1, "mineur": 2}

    changements = []
    for code in touched:
        c = constats.get(code)
        if not c:
            continue
        ap = (c.statut_valide if c.statut_valide is not None else c.statut_propose).value
        av = avant.get(code, ap)
        changements.append({
            "id": c.id,
            "code": code,
            "libelle": (libelles.get(c.exigence_ref_id, "") or "")[:160],
            "criticite": c.criticite_effective.value,
            "avant": av, "apres": ap,
            "declare": declares.get(code, ""),
            "cible": declares.get(code, "") or ap,
            "deja_valide": c.statut_valide is not None,
            "evolue": av != ap,
            "reponse": (reponses.get(code, "") or "")[:400],
        })
    changements.sort(key=lambda x: (rang.get(x["criticite"], 3), x["code"]))

    constats_list = list(constats.values())
    indic_apres = {
        "global": indicateur_global(constats_list),
        "nis2": indicateur_nis2(constats_list),
        "rgpd": indicateur_rgpd(constats_list),
        "levees": sum(1 for c in constats_list
                      if (c.statut_valide if c.statut_valide is not None else c.statut_propose) not in _EN_ATTENTE),
        "total": len(constats_list),
    }
    return {"termine": bool(dossier.analyse_termine),
            "n_reponses": len(touched),
            "n_evolutions": sum(1 for x in changements if x["evolue"]),
            "changements": changements,
            "avant": indic_avant, "apres": indic_apres}
