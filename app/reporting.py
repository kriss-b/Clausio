"""Génération des livrables : rapport PDF + fichier Excel de liaison.

- Le rapport PDF récapitule la conclusion, l'avancement, les réserves bloquantes
  et les demandes de compléments, avec les coordonnées RSSI/établissement et le
  contact du service marchés.
- Le fichier Excel de liaison porte l'identification du dossier (feuille dédiée)
  et la liste des demandes à compléter par le candidat. Réimporté via « Mise à jour
  de dossier », il est reconnu automatiquement et alimente la reprise.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from sqlmodel import Session, select

from .clausier_io import FEUILLE_DEMANDES, FEUILLE_ID
from .models import Constat, Demande, Dossier, Parametres, Preuve
from .scoring import indicateur_nis2, indicateur_rgpd, score_par_axe, statut_blocage

BASE = Path(__file__).resolve().parent.parent
EXPORTS = BASE / "exports"


def _params(session: Session) -> Parametres:
    p = session.get(Parametres, 1)
    return p or Parametres(id=1)


def _dossier_dir(dossier_id: int) -> Path:
    d = EXPORTS / str(dossier_id)
    d.mkdir(parents=True, exist_ok=True)
    return d


def _slug(s: str) -> str:
    import re
    import unicodedata
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode()
    s = re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_")
    return (s or "dossier")[:40]


def noms_fichiers(dossier: Dossier) -> dict:
    """Noms de fichiers lisibles incluant l'éditeur / le produit (non génériques)."""
    slug = _slug(dossier.societe_ou_produit or dossier.nom_affiche)
    return {"pdf": f"Clausio_rapport_{slug}.pdf",
            "excel": f"Clausio_liaison_{slug}.xlsx",
            "zip": f"Clausio_{slug}.zip"}


def _conclusion(session: Session, dossier: Dossier) -> tuple[str, dict, list[str]]:
    constats = session.exec(select(Constat).where(Constat.dossier_id == dossier.id)).all()
    non_validees = [c.code_exigence for c in constats if c.statut_valide is None]
    blocage = statut_blocage(constats)
    if non_validees:
        txt = (f"Projet de rapport — instruction incomplète : {len(non_validees)} exigence(s) "
               "restent à valider par le RSSI. Aucune conclusion d'attribution en l'état.")
    elif blocage["attribution_possible"]:
        txt = ("Avis favorable envisageable au titre de la cybersécurité, sous réserve de la "
               "levée des demandes de compléments listées.")
    else:
        txt = ("Avis défavorable en l'état : réserve(s) bloquante(s) non levée(s) — "
               f"{', '.join(blocage['bloquants_non_leves'])}. Convertible après levée des conditions.")
    return txt, blocage, non_validees


# ----------------------------------------------------------------- PDF
def _lat1(s: str) -> str:
    """Rend une chaîne compatible latin-1 (police PDF par défaut) et coupe les
    tokens trop longs pour éviter les erreurs de rendu multi_cell de fpdf2."""
    import re
    remp = {"\u2019": "'", "\u2018": "'", "\u201c": '"', "\u201d": '"',
            "\u2013": "-", "\u2014": "-", "\u2026": "...", "\u00a0": " ", "\u202f": " ",
            "\u2192": "->", "\u0153": "oe", "\u2022": "-"}
    for k, v in remp.items():
        s = s.replace(k, v)
    s = re.sub(r"(\S{45})(?=\S)", r"\1 ", s)   # coupe les mots > 45 caractères
    return s.encode("latin-1", "replace").decode("latin-1")


def generer_pdf(session: Session, dossier: Dossier) -> Path:
    from fpdf import FPDF
    from fpdf.enums import XPos, YPos

    params = _params(session)
    conclusion, blocage, non_validees = _conclusion(session, dossier)
    constats = session.exec(select(Constat).where(Constat.dossier_id == dossier.id)).all()
    demandes = session.exec(
        select(Demande).where(Demande.dossier_id == dossier.id).order_by(Demande.numero)).all()
    scores = score_par_axe(constats)
    _rang = {"bloquant": 0, "majeur": 1, "mineur": 2}
    _crit = {c.code_exigence: c.criticite_effective.value for c in constats}
    demandes = sorted(demandes, key=lambda d: (_rang.get(_crit.get(d.code_exigence, "majeur"), 3), d.code_exigence))
    nis2 = indicateur_nis2(constats)
    rgpd = indicateur_rgpd(constats)

    PRIMARY, ACCENT, INK, MUTED = (0, 0, 145), (106, 106, 244), (43, 43, 42), (110, 110, 105)
    VERDICT = {"vert": ((223, 247, 231), (24, 117, 60)),
               "orange": ((250, 238, 218), (150, 60, 0)),
               "rouge": ((252, 235, 235), (176, 0, 0))}
    if non_validees:
        verdict = "orange"
    elif blocage["attribution_possible"]:
        verdict = "vert"
    else:
        verdict = "rouge"

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.set_margins(14, 14, 14)
    pdf.add_page()
    NX, NY = XPos.LMARGIN, YPos.NEXT
    W = 210 - 28  # largeur utile

    # ---- Bandeau d'en-tete aux couleurs Clausio ----
    pdf.set_fill_color(*PRIMARY)
    pdf.rect(0, 0, 210, 30, "F")
    tx = 14
    if params.logo_path:
        lp = BASE / params.logo_path
        if lp.exists():
            try:
                pdf.image(str(lp), x=14, y=7, h=16)
                tx = 40
            except Exception:  # noqa: BLE001
                tx = 14
    pdf.set_text_color(255, 255, 255)
    pdf.set_xy(tx, 7)
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 9, "Clausio", new_x=NX, new_y=NY)
    pdf.set_x(tx)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, _lat1("Avis cybersecurite - instruction de candidature"))
    pdf.set_xy(14, 22)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(W, 5, _lat1(params.etablissement or "Etablissement de sante"), align="R")

    pdf.set_text_color(*INK)
    pdf.set_y(36)

    def section(titre):
        pdf.ln(2)
        pdf.set_text_color(*PRIMARY)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 7, _lat1(titre), new_x=NX, new_y=NY)
        y = pdf.get_y()
        pdf.set_draw_color(*ACCENT); pdf.set_line_width(0.8)
        pdf.line(14, y, 54, y)
        pdf.ln(2)
        pdf.set_text_color(*INK)

    def p(txt, size=10, style=""):
        pdf.set_font("Helvetica", style, size)
        pdf.multi_cell(0, 5, _lat1(txt), new_x=NX, new_y=NY)

    # ---- Meta dossier ----
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 6, _lat1(f"Dossier : {dossier.nom_affiche}"), new_x=NX, new_y=NY)
    pdf.set_text_color(*MUTED); pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, _lat1(f"Reference : {dossier.reference_marche}  -  Profil : {dossier.profil}  "
                         f"-  Type : {dossier.type_dispositif or 'non precise'}"), new_x=NX, new_y=NY)
    pdf.cell(0, 5, _lat1(f"Genere le {datetime.now():%d/%m/%Y a %H:%M}"), new_x=NX, new_y=NY)
    pdf.set_text_color(*INK)
    pdf.ln(3)

    # ---- Indicateurs de conformité (indicatifs) : NIS2 + RGPD ----
    def gauge(label, ind, x):
        fill, tcol = VERDICT[ind["couleur"]]
        pdf.set_xy(x, pdf.get_y())
        pdf.set_text_color(*INK); pdf.set_font("Helvetica", "B", 9)
        pdf.set_xy(x, gy); pdf.cell(88, 5, _lat1(label), new_x=NX, new_y=NY)
        by = gy + 6
        pdf.set_fill_color(225, 225, 225); pdf.rect(x, by, 60, 6, "F")
        pdf.set_fill_color(*tcol); pdf.rect(x, by, 60 * ind["score"] / 100, 6, "F")
        pdf.set_xy(x + 64, by - 2)
        pdf.set_text_color(*tcol); pdf.set_font("Helvetica", "B", 16)
        pdf.cell(24, 10, f"{ind['score']}/100", new_x=NX, new_y=NY)
        pdf.set_text_color(*MUTED); pdf.set_font("Helvetica", "", 7)
        pdf.set_xy(x, by + 8); pdf.cell(88, 4, _lat1(f"{ind['exigences']} exigences - indicatif"))

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, _lat1("Indicateurs de conformite (indicatifs, aide a la decision)"),
             new_x=NX, new_y=NY)
    gy = pdf.get_y() + 1
    gauge("Alignement NIS2 (art. 21)", nis2, 14)
    gauge("Conformite RGPD", rgpd, 112)
    pdf.set_text_color(*INK)
    pdf.set_y(gy + 20)

    # ---- Conclusion (encadre colore) ----
    section("Conclusion cybersecurite")
    fill, tcol = VERDICT[verdict]
    pdf.set_fill_color(*fill); pdf.set_draw_color(*tcol); pdf.set_line_width(0.3)
    pdf.set_text_color(*tcol); pdf.set_font("Helvetica", "B", 10)
    pdf.multi_cell(W, 6, _lat1(conclusion), border=1, fill=True, new_x=NX, new_y=NY)
    pdf.set_text_color(*INK)
    if blocage["bloquants_non_leves"]:
        pdf.ln(1)
        p("Reserves bloquantes non levees : " + ", ".join(blocage["bloquants_non_leves"]), 9, "B")

    # ---- Contacts ----
    section("Contacts")
    p(f"RSSI : {params.rssi_nom or '-'}  |  {params.rssi_email or '-'}  |  {params.rssi_tel or '-'}")
    if dossier.marche_nom or dossier.marche_email:
        p(f"Service marches : {dossier.marche_nom or '-'}  |  {dossier.marche_email or '-'}  |  {dossier.marche_tel or '-'}")
    if dossier.contact_nom or dossier.contact_email:
        p(f"Candidat : {dossier.contact_nom or '-'}  |  {dossier.contact_email or '-'}  |  {dossier.contact_tel or '-'}")

    # ---- Scores par axe ----
    section("Scores par axe (indicatifs)")
    pdf.set_font("Helvetica", "", 9)
    for axe, v in scores.items():
        sc = v["score"] if v["score"] is not None else 0
        col = (24, 117, 60) if sc >= 80 else ((150, 60, 0) if sc >= 50 else (176, 0, 0))
        pdf.cell(70, 5, _lat1(f"{axe}"))
        yb = pdf.get_y() + 1
        pdf.set_fill_color(230, 230, 230); pdf.rect(85, yb, 60, 3, "F")
        pdf.set_fill_color(*col); pdf.rect(85, yb, 60 * sc / 100, 3, "F")
        pdf.set_xy(150, pdf.get_y())
        pdf.cell(0, 5, _lat1(f"{v['score'] if v['score'] is not None else 'n/a'} / 100"), new_x=NX, new_y=NY)

    # ---- Demandes ----
    section(f"Demandes de complements ({len(demandes)})")
    if demandes:
        for d in demandes:
            pdf.set_text_color(*PRIMARY); pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(0, 4.5, _lat1(f"{d.numero}. [{d.code_exigence}]  ({d.texte})"),
                           new_x=NX, new_y=NY)
            pdf.set_text_color(*INK); pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 4.5, _lat1(d.libelle or "(intitule non disponible)"),
                           new_x=NX, new_y=NY)
            pdf.ln(1)
    else:
        p("Aucune demande de complement en attente.")

    pdf.ln(3)
    pdf.set_text_color(*MUTED); pdf.set_font("Helvetica", "I", 8)
    pdf.multi_cell(0, 4, _lat1(
        "Ce document est une pre-instruction outillee validee par le RSSI. La decision "
        "d'attribution demeure humaine et motivee. Perimetre : cybersecurite uniquement "
        "(hors dimensions financiere, fonctionnelle et juridique)."), new_x=NX, new_y=NY)

    chemin = _dossier_dir(dossier.id) / f"rapport_dossier_{dossier.id}.pdf"
    pdf.output(str(chemin))
    return chemin


# ----------------------------------------------------------------- Excel de liaison
def generer_liaison(session: Session, dossier: Dossier) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    params = _params(session)
    demandes = session.exec(
        select(Demande).where(Demande.dossier_id == dossier.id).order_by(Demande.numero)).all()
    # état actuel + criticité de chaque exigence
    constats = session.exec(select(Constat).where(Constat.dossier_id == dossier.id)).all()
    statut_par_code, crit_par_code = {}, {}
    for c in constats:
        st = c.statut_valide if c.statut_valide is not None else c.statut_propose
        statut_par_code[c.code_exigence] = st.value if st else "a_verifier"
        crit_par_code[c.code_exigence] = c.criticite_effective.value

    # tri par défaut : des plus critiques aux moins critiques (bloquant > majeur > mineur)
    _rang = {"bloquant": 0, "majeur": 1, "mineur": 2}
    demandes = sorted(demandes, key=lambda d: (_rang.get(crit_par_code.get(d.code_exigence, "majeur"), 3), d.numero))

    wb = Workbook()

    # Feuille d'identification (permet la reconnaissance automatique au réimport)
    ident = wb.active
    ident.title = FEUILLE_ID
    ident["A1"] = "NE PAS MODIFIER — identification Clausio"
    ident["A1"].font = Font(bold=True, color="B10000")
    lignes_id = [
        ("dossier_id", dossier.id),
        ("reference_marche", dossier.reference_marche),
        ("societe_ou_produit", dossier.societe_ou_produit),
        ("profil", dossier.profil),
    ]
    for i, (k, v) in enumerate(lignes_id, start=2):
        ident.cell(row=i, column=1, value=k)
        ident.cell(row=i, column=2, value=v)

    # Coordonnées (information)
    contacts = wb.create_sheet("Contacts")
    for i, (k, v) in enumerate([
        ("Établissement", params.etablissement),
        ("RSSI", params.rssi_nom), ("Courriel RSSI", params.rssi_email), ("Tél. RSSI", params.rssi_tel),
        ("Service marchés", dossier.marche_nom), ("Courriel marchés", dossier.marche_email),
        ("Tél. marchés", dossier.marche_tel),
    ], start=1):
        contacts.cell(row=i, column=1, value=k).font = Font(bold=True)
        contacts.cell(row=i, column=2, value=v)
    contacts.column_dimensions["A"].width = 22
    contacts.column_dimensions["B"].width = 40

    # Demandes à compléter par le candidat
    dem = wb.create_sheet(FEUILLE_DEMANDES)
    entetes = ["Code exigence", "Exigence à compléter", "Criticité", "Statut (à mettre à jour)",
               "Réponse du candidat", "Référence / pièce jointe"]
    for c, txt in enumerate(entetes, start=1):
        cell = dem.cell(row=1, column=c, value=txt)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="000091")
    wrap = Alignment(wrap_text=True, vertical="top")
    centre = Alignment(horizontal="center", vertical="center")
    couleur_crit = {  # fond, texte
        "bloquant": ("FFD7D7", "8B0000"),
        "majeur": ("FFE4B8", "8A5000"),
        "mineur": ("DCE6FA", "10386B"),
    }
    # tri par criticité décroissante (bloquant -> majeur -> mineur), puis par code
    _rang = {"bloquant": 0, "majeur": 1, "mineur": 2}
    demandes_triees = sorted(
        demandes, key=lambda d: (_rang.get(crit_par_code.get(d.code_exigence, "majeur"), 3), d.code_exigence))
    for i, d in enumerate(demandes_triees, start=2):
        crit = crit_par_code.get(d.code_exigence, "majeur")
        dem.cell(row=i, column=1, value=d.code_exigence).alignment = wrap
        dem.cell(row=i, column=2, value=d.libelle or "").alignment = wrap
        cc = dem.cell(row=i, column=3, value=crit)
        cc.alignment = centre
        fond, txt = couleur_crit.get(crit, ("FFFFFF", "000000"))
        cc.fill = PatternFill("solid", fgColor=fond)
        cc.font = Font(bold=True, color=txt)
        dem.cell(row=i, column=4, value=statut_par_code.get(d.code_exigence, "a_verifier")).alignment = wrap
        dem.cell(row=i, column=5, value="").alignment = wrap
        dem.cell(row=i, column=6, value="").alignment = wrap

    # Menu déroulant sur la colonne « Statut » (colonne D)
    if demandes:
        dv = DataValidation(
            type="list",
            formula1='"couvert,partiel,absent,non_applicable,a_verifier"',
            allow_blank=True, showDropDown=False)
        dv.promptTitle = "Statut de l'exigence"
        dv.prompt = "Sélectionnez l'état, puis renseignez votre réponse dans la colonne suivante."
        dem.add_data_validation(dv)
        dv.add(f"D2:D{len(demandes) + 1}")

    dem.column_dimensions["A"].width = 14
    dem.column_dimensions["B"].width = 70
    dem.column_dimensions["C"].width = 12
    dem.column_dimensions["D"].width = 20
    dem.column_dimensions["E"].width = 48
    dem.column_dimensions["F"].width = 28
    dem.freeze_panes = "A2"

    chemin = _dossier_dir(dossier.id) / f"liaison_dossier_{dossier.id}.xlsx"
    wb.save(chemin)
    return chemin


def generer_zip(session: Session, dossier: Dossier, pdf: Path, xls: Path) -> Path:
    import zipfile
    noms = noms_fichiers(dossier)
    chemin = _dossier_dir(dossier.id) / f"clausio_dossier_{dossier.id}.zip"
    with zipfile.ZipFile(chemin, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(pdf, arcname=noms["pdf"])
        z.write(xls, arcname=noms["excel"])
    return chemin


def generer_livrables(session: Session, dossier: Dossier) -> dict:
    conclusion, blocage, non_validees = _conclusion(session, dossier)
    constats = session.exec(select(Constat).where(Constat.dossier_id == dossier.id)).all()
    pdf = generer_pdf(session, dossier)
    xls = generer_liaison(session, dossier)
    zip_ = generer_zip(session, dossier, pdf, xls)
    return {
        "conclusion": conclusion,
        "instruction_complete": not non_validees,
        "exigences_en_attente_rssi": non_validees,
        "blocage": blocage,
        "nis2": indicateur_nis2(constats),
        "rgpd": indicateur_rgpd(constats),
        "pdf": pdf.name,
        "excel": xls.name,
        "zip": zip_.name,
    }
